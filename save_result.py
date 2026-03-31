import pandas as pd
import os
import warnings
from packaging import version

def save_result(filename, head, data, sheet_name, folder_name):
    from openpyxl import load_workbook

    df_new = pd.DataFrame(data, columns=head)
    file_exists = os.path.exists(filename)

    # 判断是否支持if_sheet_exists参数（pandas >= 1.3.0），否则先删除表
    support_if_sheet_exists = version.parse(pd.__version__) >= version.parse("1.3.0")

    if sheet_name == "summary" and os.path.exists(filename) and support_if_sheet_exists:
        # summary需要追加，而不是直接replace
        try:
            df_exist = pd.read_excel(filename, sheet_name="summary")
            df_new = pd.DataFrame(data, columns=head)
            df_total = pd.concat([df_exist, df_new], ignore_index=True)
        except Exception:
            # 没有summary表时新建
            df_total = pd.DataFrame(data, columns=head)
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_total.to_excel(writer, sheet_name="summary", index=False)
        return

    if file_exists:
        # 检查sheet是否存在
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            # 新pandas直接replace，老pandas先删再写
            if support_if_sheet_exists:
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # 老pandas不支持replace，先删除sheet再保存
                del wb[sheet_name]
                wb.save(filename)
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                    df_new.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # sheet不存在时正常添加
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # 文件不存在直接新建
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df_new.to_excel(writer, sheet_name=sheet_name, index=False)